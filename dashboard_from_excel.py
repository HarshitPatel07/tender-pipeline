"""Build the HTML dashboard from an already-produced Tender_Summary.xlsx.

Use this when you have the Excel but do not want to re-read all the PDFs:
    python dashboard_from_excel.py "Tender_Summary.xlsx" [out.html]
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import openpyxl
import tender_extractor as te

src = Path(sys.argv[1] if len(sys.argv) > 1 else "Tender_Summary.xlsx")
dst = Path(sys.argv[2] if len(sys.argv) > 2 else "Tender_Dashboard.html")

wb = openpyxl.load_workbook(src)
label_to_key = {label: key for key, label in te.FIELDS}

# --- Evidence sheet carries source ref, confidence and flag per field --------
ev = {}
if "Evidence" in wb.sheetnames:
    for r in wb["Evidence"].iter_rows(min_row=2, values_only=True):
        tender, field, val, ref, conf, rules, ai, flag = (list(r) + [None] * 8)[:8]
        if not tender or field not in label_to_key:
            continue
        ev.setdefault(tender, {})[label_to_key[field]] = te.Result(
            value=str(val or te.NOT_FOUND), ref=str(ref or ""),
            conf=str(conf or ""), rules_value=str(rules or ""),
            ai_value=str(ai or ""), flag=str(flag or ""))

# --- Documents Read sheet ----------------------------------------------------
docs = {}
if "Documents Read" in wb.sheetnames:
    for r in wb["Documents Read"].iter_rows(min_row=2, values_only=True):
        tender, name, pages, ocr, note = (list(r) + [None] * 5)[:5]
        if not tender or not name:
            continue
        docs.setdefault(tender, []).append(
            {"name": str(name), "pages": int(pages or 0),
             "ocr": int(ocr or 0), "note": str(note or "")})

# --- Summary sheet gives the row order and the final values -----------------
ws = wb["Tender Summary"]
headers = [c.value for c in ws[1]]
rows = []
for r in ws.iter_rows(min_row=2):
    folder = r[0].value
    if not folder:
        continue
    results = {}
    for i, h in enumerate(headers):
        key = label_to_key.get(h)
        if not key:
            continue
        cached = ev.get(folder, {}).get(key)
        results[key] = cached or te.Result(value=str(r[i].value or te.NOT_FOUND))
    for key, _ in te.FIELDS:
        results.setdefault(key, te.Result())
    rows.append({"tender": folder, "results": results,
                 "files": docs.get(folder, []), "ai_ok": True})

out = te.write_dashboard(rows, dst, stamp=f"From {src.name}")
flagged = sum(1 for x in rows for k, _ in te.FIELDS if x["results"][k].flag)
print(f"{len(rows)} tenders, {flagged} flagged field(s) -> {out}")
